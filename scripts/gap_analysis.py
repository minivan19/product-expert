# -*- coding: utf-8 -*-
"""
gap_analysis.py - 客户功能缺口分析
三步分析框架：
  Step 1: 买了没？    → 合同/主数据 + 关键词匹配
  Step 2: 实施了没？  → 蓝图方案 + 内容关键词匹配
  Step 3: 用了没？    → 运维工单 + LLM凝练
基准：references/product_modules_hierarchy.json
"""

import os
import re
import json
import sys
import openpyxl
import win32com.client
import pythoncom
import zipfile
import fitz
import unicodedata

CLIENT_DATA_ROOT = r"C:\Users\mingh\client-data\raw\客户档案"
HIERARCHY_PATH = os.path.join(os.path.dirname(__file__), "..", "references", "product_modules_hierarchy.json")

# 延迟导入 term_map（避免循环依赖）
def _get_term_map():
    import term_map as tm
    return tm


# -----------------------------------------
# 工具函数
# -----------------------------------------

def norm(s: str) -> str:
    """消除全角ASCII/标点，CJK字符不变"""
    return (s
        .replace('\uff08', '(').replace('\uff09', ')')
        .replace('\uff0f', '/')
        .replace('\uff0c', ',').replace('\uff0e', '.').replace('\uff1b', ';').replace('\uff1a', ':')
    )


def load_hierarchy() -> list:
    """加载产品模块层级JSON"""
    with open(HIERARCHY_PATH, encoding='utf-8') as f:
        return json.load(f)


def build_module_kw_map(hierarchy: list) -> dict:
    """
    为每个模块构建关键词集合（用于蓝图/合同匹配）
    key = 模块名（去序号前缀）
    value = {module: str, features: set, all_keywords: set}
    """
    mod_map = {}
    for item in hierarchy:
        mod_raw = item['module']
        # 去掉序号前缀："1.基础供应商管理" → "基础供应商管理"
        if mod_raw[0].isdigit() and '.' in mod_raw[:3]:
            mod_name = mod_raw.split('.', 1)[1]
        else:
            mod_name = mod_raw

        kws = set()
        kws.add(mod_name)
        if mod_name not in kws:
            kws.add(mod_name)
        for feat in item.get('features', []):
            kws.add(feat)

        mod_map[mod_name] = {
            'module': mod_raw,
            'suite': item.get('suite', ''),
            'features': set(item.get('features', [])),
            'all_keywords': kws
        }
    return mod_map


# Excel购买模块名 → 产品模块名 的映射（已确认诺斯贝尔的命名差异）
EXCEL_TO_PRODUCT = {
    '基础供应商管理': '基础供应商管理',
    '寻源管理': '基础采购寻源',
    '基础采购协同': '基础采购协同',
    '商城采购': '商城采购',
    '自有供应商目录化': '自有供应商目录化',
    '高级供应商管理': '高级供应商管理',
    '供应商管理': '基础供应商管理',  # 部分匹配
}


# -----------------------------------------
# Step 1: 买了没？（合同优先 → 主数据备选）
# -----------------------------------------

def read_bought_from_master(client_dir: str) -> set:
    """
    从客户主数据xlsx读取"购买模块"列，返回产品模块名集合
    """
    master_path = os.path.join(client_dir, '基础数据')
    if not os.path.isdir(master_path):
        return set()

    xlsx_files = [f for f in os.listdir(master_path) if f.endswith('.xlsx') and not f.startswith('~$')]
    for fn in xlsx_files:
        if '客户主数据' in fn or '主数据' in fn:
            fp = os.path.join(master_path, fn)
            break
    else:
        return set()

    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map = {str(h).strip() if h else '': i for i, h in enumerate(row1)}

    # 找"购买模块"或"产品模块"列
    mod_col = None
    for name, idx in col_map.items():
        if '购买模块' in name or '产品模块' in name:
            mod_col = idx
            break

    if mod_col is None:
        wb.close()
        return set()

    bought = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[mod_col]:
            raw = str(row[mod_col]).strip()
            if raw and raw not in ('无', '空', 'None'):
                # 处理"基础供应商管理,寻源管理,..."格式
                for part in raw.split(','):
                    part = part.strip()
                    if not part:
                        continue
                    # 映射到产品模块名
                    product_mod = EXCEL_TO_PRODUCT.get(part, part)
                    bought.add(product_mod)

    wb.close()
    return bought


def read_bought_from_contracts(client_dir: str) -> set:
    """
    从订阅合同明细xlsx读取合同产品名，返回产品模块名集合
    合同数据是产品级，不是模块级，所以要做额外解析
    """
    sub_path = os.path.join(client_dir, '订阅合同行')
    if not os.path.isdir(sub_path):
        return set()

    xlsx_files = [f for f in os.listdir(sub_path) if f.endswith('.xlsx') and not f.startswith('~$')]
    found = None
    for fn in xlsx_files:
        if '明细' in fn or '订阅' in fn:
            found = os.path.join(sub_path, fn)
            break

    if not found:
        return set()

    wb = openpyxl.load_workbook(found, data_only=True)
    ws = wb.active
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map = {str(h).strip() if h else '': i for i, h in enumerate(row1)}

    # 产品名称列
    prod_col = col_map.get('产品名称', None)
    if prod_col is None:
        wb.close()
        return set()

    bought = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[prod_col]:
            raw = str(row[prod_col]).strip()
            if raw and raw not in ('无', '空', 'None'):
                # 合同中的产品名通常包含SRM/供应商/寻源等关键词
                # 这里根据关键词推断模块（合同级，无法精确到模块）
                if '供应商' in raw: bought.add('基础供应商管理')
                if '寻源' in raw: bought.add('基础采购寻源')
                if '采购' in raw and '协同' in raw: bought.add('基础采购协同')
                if '商城' in raw: bought.add('商城采购')
                if '目录' in raw: bought.add('自有供应商目录化')

    wb.close()
    return bought


def step1_bought_modules(client_dir: str) -> dict:
    """
    Step 1: 买了哪些模块？
    优先从合同读取，其次从主数据
    返回: {产品模块名: True/False}
    """
    hierarchy = load_hierarchy()
    mod_names = {m['module'].split('.', 1)[1] if '.' in m['module'][:3] else m['module']
                 for m in hierarchy}

    # 优先合同
    contract_bought = read_bought_from_contracts(client_dir)
    master_bought = read_bought_from_master(client_dir)

    # 合并：合同优先，合同没有的用主数据补充
    bought = contract_bought | master_bought

    result = {}
    for mod in sorted(mod_names):
        result[mod] = mod in bought

    print(f"  [Step1 买了没] 合同:{len(contract_bought)} 主数据:{len(master_bought)} → 合并:{len(bought)}")
    for mod, has in sorted(result.items()):
        if has:
            print(f"    [Y] {mod}")
        else:
            print(f"    [N] {mod}")
    return result


# -----------------------------------------
# Step 2: 实施了没？（蓝图方案内容匹配）
# -----------------------------------------

def extract_doc_text(fp: str) -> str:
    """读取DOC/DOCX正文（段落 + 表格内容）"""
    text_parts = []
    word, doc = None, None
    pythoncom.CoInitialize()
    try:
        word = win32com.client.Dispatch('Word.Application')
        word.Visible = False
        word.DisplayAlerts = False
        doc = word.Documents.Open(os.path.abspath(fp), ReadOnly=True, ConfirmConversions=False)
        # 读段落
        for para in doc.Paragraphs:
            t = para.Range.Text.strip()
            if t and len(t) > 1:
                text_parts.append(t)
        # 读表格（诺斯贝尔DOC的操作内容在表格里）
        if doc.Tables.Count > 0:
            for tbl in doc.Tables:
                for row in tbl.Rows:
                    for cell in row.Cells:
                        t = cell.Range.Text.strip()
                        if t and len(t) > 1:
                            text_parts.append(t)
    except Exception as e:
        text_parts.append(f"[DOC错误: {e}]")
    finally:
        if doc:
            try: doc.Close(False)
            except: pass
        if word:
            try: word.Quit()
            except: pass
        del word, doc
    pythoncom.CoUninitialize()
    return '\n'.join(text_parts)


def extract_pptx_text(fp: str) -> str:
    """读取PPTX/PPT正文（ZIP+XML方式）"""
    text_parts = []
    try:
        with zipfile.ZipFile(fp) as z:
            for name in z.namelist():
                if re.match(r'ppt/slides/slide\d+\.xml', name):
                    with z.open(name) as f:
                        content = f.read().decode('utf-8', errors='replace')
                        texts = re.findall(r'<a:t[^>]*>([^<]+)</a:t>', content)
                        for t in texts:
                            t = t.strip()
                            if t:
                                text_parts.append(t)
    except Exception as e:
        return f"[PPTX错误: {e}]"
    return '\n'.join(text_parts)


def extract_blueprint(fp: str) -> str:
    """统一入口：根据扩展名调用不同提取方法"""
    ext = os.path.splitext(fp)[1].lower()
    if ext in ('.doc', '.docx'):
        return extract_doc_text(fp)
    elif ext in ('.pptx', '.ppt'):
        return extract_pptx_text(fp)
    elif ext == '.pdf':
        # PDF流程图：正文提取失败，用文件名降级
        name = os.path.splitext(os.path.basename(fp))[0]
        return name
    return ""


def step2_implemented_modules(client_dir: str) -> dict:
    """
    Step 2: 哪些模块在蓝图中被实施？
    遍历蓝图文件，用模块/功能的关键词匹配内容
    返回: {产品模块名: {'implemented': set(), 'files': [file_list]}}
    """
    blueprint_dir = os.path.join(client_dir, '蓝图方案')
    if not os.path.isdir(blueprint_dir):
        print("  [Step2 实施了没] 蓝图文件夹不存在")
        return {}

    hierarchy = load_hierarchy()
    mod_kw_map = build_module_kw_map(hierarchy)

    # 每个模块追踪：实现了哪些功能
    impl = {mod: {'implemented': set(), 'files': set()} for mod in mod_kw_map}

    files = sorted(os.listdir(blueprint_dir))
    print(f"  [Step2 实施了没] 扫描蓝图文件: {len(files)}个")

    for fn in files:
        fp = os.path.join(blueprint_dir, fn)
        text = extract_blueprint(fp)
        text_norm = norm(text)

        for mod_name, mod_info in mod_kw_map.items():
            kws = mod_info['all_keywords']
            hit_feats = []
            for kw in kws:
                if norm(kw) in text_norm:
                    hit_feats.append(kw)
            # 兜底：如果精确关键词全没匹配，用模块名单字短词匹配
            if not hit_feats and len(mod_name) >= 3:
                short_kws = [c for c in ['商城', '采购', '寻源', '询价', '供应商',
                                          '协同', '目录', '会员', '预算', '合同',
                                          '资质', '绩效', '招标', '投标', '竞价',
                                          '结算', '付款', '质量', '库存', '移动']
                           if c in mod_name]
                for kw in short_kws:
                    if norm(kw) in text_norm:
                        hit_feats.append(f'[短词]{kw}')
            if hit_feats:
                impl[mod_name]['files'].add(fn)
                for feat in hit_feats:
                    # 精确关键词: 必须在 features 里才算implemented
                    # 短词标记: 不在 features 里，但也记录（说明有相关蓝图）
                    if feat.startswith('[短词]') or feat in mod_info['features']:
                        impl[mod_name]['implemented'].add(feat)

    # 汇总输出
    for mod_name, data in sorted(impl.items()):
        total_feats = len(mod_kw_map[mod_name]['features'])
        impl_count = len(data['implemented'])
        status = f"{impl_count}/{total_feats}" if total_feats > 0 else "?"
        print(f"    {'[Y]' if data['files'] else '[N]'} {mod_name}: {status}功能, {len(data['files'])}个文件")

    return impl


# -----------------------------------------
# Step 3: 用了没？（运维工单 + LLM）
# -----------------------------------------

def _read_workorder_batch(fp, year):
    """读取单个工单xlsx"""
    records = []
    try:
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        col_map = {}
        for i, h in enumerate(row1):
            if h:
                col_map[str(h).strip()] = i
        tc = col_map.get('标题', col_map.get('工单号', 1))
        dc = col_map.get('描述', col_map.get('详细', col_map.get('问题描述', 9)))
        mod_i = col_map.get('模块', 7)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            title = str(row[tc]).strip() if row[tc] else ''
            desc = str(row[dc]).strip() if row[dc] else ''
            mod = str(row[mod_i]).strip() if row[mod_i] else ''
            if title or desc:
                records.append({"模块": mod, "标题": title, "描述": desc[:300]})
        wb.close()
    except Exception:
        pass
    return records


def read_workorders(client_dir: str, year: int) -> list:
    """读取所有工单记录"""
    ops_dir = os.path.join(client_dir, '运维工单')
    if not os.path.isdir(ops_dir):
        return []

    all_records = []
    files = sorted([f for f in os.listdir(ops_dir) if f.endswith('.xlsx')])
    for fn in files:
        all_records.extend(_read_workorder_batch(os.path.join(ops_dir, fn), year))
    return all_records


def step3_used_modules(client_dir: str, year: int = 2025) -> dict:
    """
    Step 3: 工单中使用了哪些模块？（术语映射+LLM提取）
    返回: {产品模块名: count}
    """
    tm = _get_term_map()
    records = read_workorders(client_dir, year)
    print(f"  [Step3 用了没] 工单记录: {len(records)}条")

    # 调用 term_map 做分析（interactive=False，静默模式）
    feature_counts = tm.analyze_workorders(records, interactive=False)

    # 聚合到模块级别（用于3×2分类）
    hierarchy = load_hierarchy()
    mod_kw_map = build_module_kw_map(hierarchy)
    usage = {mod: 0 for mod in mod_kw_map}
    for module, features in feature_counts.items():
        if module in usage:
            usage[module] = sum(features.values())

    for mod, cnt in sorted(usage.items(), key=lambda x: -x[1]):
        bar = '#' * min(cnt, 20)
        print(f"    {'[Y]' if cnt > 0 else '·'} {mod}: {cnt}次 {bar}")

    # 返回两个 dict：module_usage（模块级count） + feature_counts（功能级详情）
    return {'module': usage, 'feature': feature_counts}


# -----------------------------------------
# 三步综合 → 3×2 网格分类
# -----------------------------------------

def classify_3x2(bought: dict, implemented: dict, used: dict, hierarchy: list) -> dict:
    """
    综合三步结果，输出3×2网格分类
    维度1: 买了没（bought）
    维度2: 用了多少（used count）

    A=买了+深度使用(>5)  B=买了+轻度使用(1-5)  C=买了+未使用(0)
    D=没买+工单有        E=没买+工单无
    """
    mod_names = [m['module'].split('.', 1)[1] if '.' in m['module'][:3] else m['module']
                 for m in hierarchy]

    result = {'A': [], 'B': [], 'C': [], 'D': [], 'E': []}

    for mod in sorted(mod_names):
        has_bought = bought.get(mod, False)
        cnt = used.get(mod, 0)

        if has_bought:
            if cnt > 5:
                result['A'].append(mod)
            elif cnt > 0:
                result['B'].append(mod)
            else:
                result['C'].append(mod)
        else:
            if cnt > 0:
                result['D'].append(mod)
            else:
                result['E'].append(mod)

    print(f"\n  [3×2分类结果]")
    labels = {'A': '深度应用(买了+高频>5)', 'B': '激活不足(买了+低频1-5)',
               'C': '买了未实施(买了+0)', 'D': '潜在需求(没买+工单有)', 'E': '空白机会(没买+工单无)'}
    for cls in 'ABCDE':
        mods = result[cls]
        print(f"    [{cls}] {labels[cls]}: {len(mods)}个 {' '.join(mods[:3])}{'...' if len(mods)>3 else ''}")

    return result


# -----------------------------------------
# LLM 推荐生成（Phase A/B/C/D）
# -----------------------------------------

def call_llm(messages: list, model: str = None) -> str:
    """调用LLM接口（优先DeepSeek，fallback MiniMax）"""
    try:
        from openai import OpenAI
    except ImportError:
        return "[LLM未安装]"

    # 优先用 DeepSeek（从openclaw.json读取）
    deepseek_cfg = None
    try:
        import json
        oc_path = os.path.join(os.path.expanduser('~'), '.openclaw', 'openclaw.json')
        with open(oc_path, encoding='utf-8') as f:
            data = json.load(f)
        providers = data.get('models', {}).get('providers', {})
        for name, cfg in providers.items():
            if 'deepseek' in name.lower():
                deepseek_cfg = cfg
                break
    except Exception:
        pass

    if deepseek_cfg:
        client = OpenAI(
            api_key=deepseek_cfg.get('apiKey', ''),
            base_url=deepseek_cfg.get('baseUrl', '')
        )
        model_id = model
        if not model_id:
            models = deepseek_cfg.get('models', [])
            model_id = models[0].get('id', 'deepseek-chat') if models else 'deepseek-chat'
    else:
        client = OpenAI(
            api_key=os.environ.get('OPENAI_API_KEY', ''),
            base_url=os.environ.get('OPENAI_API_BASE', 'https://api.minimaxi.com/v1')
        )
        model_id = model or 'MiniMax-Accelerate'

    try:
        resp = client.chat.completions.create(
            model=model_id,
            messages=messages,
            temperature=0.7,
            max_tokens=2000
        )
        return resp.choices[0].message.content
    except Exception as e:
        return f"[LLM错误: {e}]"


def _read_blueprint_for_module(impl: dict, mod: str, client_dir: str) -> str:
    """为C类分析读取蓝图文本：提取指定模块相关的蓝图内容"""
    blueprint_dir = os.path.join(client_dir, '蓝图方案')
    if not os.path.isdir(blueprint_dir):
        return ""
    files = impl.get(mod, {}).get('files', set())
    if not files:
        return ""
    texts = []
    for fn in files:
        fp = os.path.join(blueprint_dir, fn)
        if os.path.exists(fp):
            txt = extract_blueprint(fp)
            if txt and len(txt) > 20:
                texts.append(f"【{fn}】\n{txt[:800]}")
    return '\n---\n'.join(texts)


def _get_module_features(hierarchy: list, mod_name: str) -> list:
    """从hierarchy中获取指定模块的所有功能名"""
    for item in hierarchy:
        mod_raw = item['module']
        if mod_raw[0].isdigit() and '.' in mod_raw[:3]:
            name = mod_raw.split('.', 1)[1]
        else:
            name = mod_raw
        if name == mod_name:
            return item.get('features', [])
    return []


def _qdrant_search(query: str, top_k: int = 5) -> list:
    """Qdrant搜索产品功能"""
    try:
        sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
        from src.qdrant_ops import search_points
        return search_points(query, top_k=top_k)
    except Exception as e:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        print(f"    [Qdrant搜索失败] {e}")
        return []


def _format_qdrant_results(results: list) -> list:
    """整理Qdrant结果"""
    formatted = []
    for payload, score in results:
        if not payload:
            continue
        text = payload.get('text', '')[:300]
        module = payload.get('module', '')
        doc = payload.get('doc_name', '')
        formatted.append({
            'module': module,
            'text': text,
            'doc': doc,
            'score': round(score, 3)
        })
    return formatted


def _summarize_qdrant(qdrant_raw: list, mods: list, client_name: str) -> list:
    """
    对Qdrant原始结果去重 + LLM总结，输出标准化结构。
    输出每条：{module, feature, usage, value}
    """
    if not qdrant_raw:
        return []

    # --- 去重：长文本Exact dedup -----------------------------------------
    seen_texts = set()
    deduped = []
    for item in qdrant_raw:
        text = item.get('text', '')
        if len(text) > 50:
            key = text[:80]
            if key not in seen_texts:
                seen_texts.add(key)
                deduped.append(item)
        else:
            deduped.append(item)  # 短文本（路径节点）直接保留

    # --- 截断总量上限，避免单次prompt过长 ---------------------------
    MAX_ITEMS = 30
    deduped = deduped[:MAX_ITEMS]

    # --- 构建LLM batch summarization prompt ----------------------------
    items_text = []
    for i, item in enumerate(deduped, 1):
        items_text.append(
            f"[{i}] 模块:{item['module']} | 文本:{item['text'][:200]}"
        )
    items_block = '\n'.join(items_text)

    mod_list = ' / '.join(mods) if mods else '未知'

    prompt = f"""你是SRM产品功能知识库管理员。客户「{client_name}」的模块：{mod_list}。

参考以下从产品知识库检索到的内容，为每一条生成标准化推荐摘要：

{items_block}

要求：
- 每条输出4个字段：功能名称（从文本中抽取，不超过15字）、功能用途（≤20字）、业务价值（≤20字）
- 功能名称要具体，不能模糊（如"订单管理"而不是"功能模块"）
- 同一功能的重复条目只保留1条
- 不要编造信息，基于原文推理

输出格式（纯JSON数组，每条4个字段）：
[
  {{"feature": "xxx", "usage": "xxx", "value": "xxx"}},
  ...
]
只输出JSON，不要其他文字。"""

    raw = call_llm([{"role": "user", "content": prompt}])

    # --- 解析JSON -------------------------------------------------------
    try:
        import json, re
        # 尝试直接解析
        try:
            data = json.loads(raw)
        except Exception:
            # 尝试从 ```json ... ``` 中提取
            m = re.search(r'\[.*\]', raw, re.DOTALL)
            if m:
                data = json.loads(m.group(0))
            else:
                data = []
    except Exception:
        data = []

    # --- 组装结果，附加module信息 ----------------------------------------
    results = []
    for item in data:
        if isinstance(item, dict) and item.get('feature'):
            results.append({
                'module': item.get('module', ''),
                'feature': item.get('feature', ''),
                'usage': item.get('usage', ''),
                'value': item.get('value', ''),
            })
    return results


def generate_recommendations(grid: dict, used: dict, client_name: str,
                              impl: dict = None, hierarchy: list = None,
                              client_dir: str = None,
                              feature_counts: dict = None) -> dict:
    """
    为各分类生成推荐内容（分层策略）：
    A类：同模块找未深度使用的亮点功能
    B类：同功能找更多使用场景
    C类：蓝图分析 → 预测痛点 → Qdrant精准匹配
    D类：基于工单内容推荐
    """
    labels = {
        'A': '深度应用模块',
        'B': '激活不足模块',
        'C': '疑似未用模块',
        'D': '潜在需求模块',
    }
    results = {}

    for cls in 'ABCD':
        if not grid.get(cls):
            results[cls] = {}
            continue

        mods = grid[cls]
        raw = ""
        qdrant = []

        if cls == 'A':
            # 同模块→高级模块映射（基于hierarchy的suite分组）
            MODULE_UPGRADE = {
                '基础供应商管理': '高级供应商管理',
                '基础采购协同': '高级采购协同',
            }
            for mod in mods:
                all_feats = _get_module_features(hierarchy, mod)
                used_feats = (feature_counts or {}).get(mod, {})  # {feature: count}
                underused, unused_in_mod = [], []
                for feat, cnt in used_feats.items():
                    if 1 <= cnt <= 5:
                        underused.append((feat, cnt))
                for feat in all_feats:
                    if feat not in used_feats:
                        unused_in_mod.append(feat)

                upgrade_target = MODULE_UPGRADE.get(mod, None)
                upgrade_feats = _get_module_features(hierarchy, upgrade_target) if upgrade_target else []
                used_summary = {f: c for f, c in used_feats.items() if c > 0}

                prompt = f"""你是SRM客户成功顾问。客户「{client_name}」的「{mod}」模块已深度应用（高频使用）。

已深度使用的功能：{used_summary}
同模块内未充分使用的功能：{underused}
同模块内完全未使用的功能：{unused_in_mod}
{('该模块有升级版本【' + upgrade_target + '】，其核心功能：' + str(upgrade_feats)) if upgrade_target else '（无直接升级版本）'}

请从以下两个维度给出推荐：

【维度一：同模块深挖】
在「{mod}」模块内部：
1. 分析：为什么客户在这些功能上用得浅或没用？
2. 选出2个最有激活价值的"亮点功能"（优先从未使用中选择，其次选低频使用的）
3. 对每个亮点功能，给出1条具体的激活建议（15字以内，要包含具体功能名）

【维度二：跨模块升级】（如果有升级版本）
如果「{mod}」有对应的「{upgrade_target or '更高级版本'}」，分析：
1. 该客户是否适合升级？（基于其当前使用深度）
2. 升级后最值得优先启用的2个高级功能是什么？
3. 给出一条升级路径建议（15字以内）

回复格式：
【同模块深挖】
分析：...
亮点1：[功能名] → [激活建议]
亮点2：[功能名] → [激活建议]
【跨模块升级】
{"分析：... | 升级功能1：[功能名] → [升级建议] | 升级功能2：[功能名] → [升级建议]" if upgrade_target else "（无升级版本，此维度略过）"}
"""
                raw_mod = call_llm([{"role": "user", "content": prompt}])
                raw += f"【{mod}】\n{raw_mod}\n\n"

                # Qdrant：两个方向都搜
                same_query = f"{mod} {' '.join(unused_in_mod[:3])}"
                qdrant_same = _qdrant_search(same_query, top_k=4)
                qdrant.extend(_format_qdrant_results(qdrant_same))
                if upgrade_target:
                    upgrade_query = f"{upgrade_target} {''.join(upgrade_feats[:3])}"
                    qdrant_up = _qdrant_search(upgrade_query, top_k=4)
                    qdrant.extend(_format_qdrant_results(qdrant_up))

        elif cls == 'B':
            # ===== B类：激活不足 → 同功能找更多使用场景 =====
            for mod in mods:
                used_feats = (feature_counts or {}).get(mod, {})  # {feature: count}
                if not used_feats:
                    continue

                # 找出该模块下所有已使用的功能及其频次
                active_feats = [(f, c) for f, c in used_feats.items() if c > 0]
                all_feats = _get_module_features(hierarchy, mod)
                other_feats = [f for f in all_feats if f not in used_feats]

                prompt = f"""你是SRM客户成功顾问。客户「{client_name}」的「{mod}」模块使用频率较低。

已使用的功能及频次：{dict(active_feats)}
同模块其他可用功能：{other_feats}

请完成：
1. 分析：为什么「{mod}」整体使用频率不高？可能原因有哪些？
2. 针对已使用的每个功能，给出1条"深度激活"建议（15字以内，要包含该功能的更高级用法或关联场景）
3. 针对同模块其他功能，选取2个最有价值的，给出激活路径建议（15字以内）

回复格式：
分析：[50字以内分析]
深度激活：[功能1] → [更高级用法] | [功能2] → [更高级用法]
模块扩展：[功能X] → [激活路径] | [功能Y] → [激活路径]
"""
                raw_mod = call_llm([{"role": "user", "content": prompt}])
                raw += f"【{mod}】\n{raw_mod}\n\n"

                # Qdrant：搜索该模块高级用法
                qdrant_mod = _qdrant_search(f"{mod} 高级用法 {''.join([f for f,_ in active_feats[:2]])}", top_k=5)
                qdrant.extend(_format_qdrant_results(qdrant_mod))

        elif cls == 'C':
            # ===== C类：疑似未用 → 直接推荐模块内未实施功能 =====
            for mod in mods:
                # 找出该模块买了但未实施的功能（来自Step2的覆盖率数据）
                impl_mod = (impl or {}).get(mod, {})
                total_feats = len(_get_module_features(hierarchy, mod))
                impl_feats = impl_mod.get('implemented', set())  # 已实施的功能名
                all_feats = set(_get_module_features(hierarchy, mod))
                unimplemented = [f for f in all_feats if f not in impl_feats]

                # 无可用功能清单时，用蓝图文本辅助
                blueprint_text = _read_blueprint_for_module(impl or {}, mod, client_dir or "")
                feats_str = '、'.join(unimplemented) if unimplemented else '（未知具体功能）'

                prompt = f"""你是SRM客户成功顾问。客户「{client_name}」购买了「{mod}」模块，但几乎没用过。

该模块已知的未实施功能：{feats_str}
{'客户蓝图摘要：' + blueprint_text[:500] if blueprint_text else '（无蓝图文件）'}

请完成：
1. 分析：该模块"买而未用"最可能的原因（≤40字）
2. 针对未实施功能，选出2个优先推荐的功能，给出激活建议（15字以内，含具体功能名）

回复格式：
分析：...
推荐1：[功能名] → [激活建议]
推荐2：[功能名] → [激活建议]
"""
                raw_mod = call_llm([{"role": "user", "content": prompt}])
                raw += f"【{mod}】\n{raw_mod}\n\n"

                # Qdrant：搜未实施功能的产品知识
                search_kw = ' '.join(unimplemented[:4]) if unimplemented else mod
                qdrant_mod = _qdrant_search(f"{mod} {search_kw}", top_k=6)
                qdrant.extend(_format_qdrant_results(qdrant_mod))

        elif cls == 'D':
            # ===== D类：潜在需求 → 基于工单内容推荐 =====
            mod_list = ', '.join(mods)
            prompt = f"""你是SRM客户成功顾问。客户「{client_name}」未购买以下模块，但运维工单中出现了相关术语：{mod_list}

请基于SRM行业经验，给出2-3个模块的优先级推荐（说明为什么这些模块值得购买）。

回复格式：
模块1（优先级高）：推荐理由
模块2（优先级中）：推荐理由
"""
            raw = call_llm([{"role": "user", "content": prompt}])
            for mod in mods:
                qdrant_mod = _qdrant_search(f"{mod} 核心功能 适用场景", top_k=4)
                qdrant.extend(_format_qdrant_results(qdrant_mod))

        results[cls] = {
            'raw': raw.strip(),
            'mods': grid[cls],
            'qdrant': _summarize_qdrant(qdrant, grid[cls], client_name)
        }
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        print(f"    [推荐] {cls}类: {len(qdrant)}条Qdrant结果")

    return results


# -----------------------------------------
# 主入口
# -----------------------------------------

def find_client_dir(client_name: str) -> str:
    """查找客户数据目录"""
    if not os.path.isdir(CLIENT_DATA_ROOT):
        raise FileNotFoundError(f"客户档案目录不存在: {CLIENT_DATA_ROOT}")

    for name in os.listdir(CLIENT_DATA_ROOT):
        if client_name in name:
            return os.path.join(CLIENT_DATA_ROOT, name)
    raise FileNotFoundError(f"未找到客户目录: {client_name}")


def main(client_name: str, year: int = 2025, output_path: str = None):
    """三步分析主流程"""
    print(f"\n{'='*50}")
    print(f"客户: {client_name} | 年份: {year}")
    print(f"{'='*50}")

    hierarchy = load_hierarchy()
    client_dir = find_client_dir(client_name)

    # Step 1: 买了没
    print(f"\n[Step1] 买了哪些模块？")
    bought = step1_bought_modules(client_dir)

    # Step 2: 实施了没
    print(f"\n[Step2] 哪些模块已实施？")
    impl = step2_implemented_modules(client_dir)

    # Step 3: 用了没
    print(f"\n[Step3] 工单中使用情况？")
    used_data = step3_used_modules(client_dir, year)
    used = used_data['module']          # {module: count}，用于3×2分类
    feature_counts = used_data['feature']  # {module: {feature: count}}，用于推荐详情

    # 3×2 分类
    print(f"\n[综合] 3×2网格分类")
    grid = classify_3x2(bought, impl, used, hierarchy)

    # 生成推荐
    print(f"\n[推荐] 生成激活建议...")
    recs = generate_recommendations(grid, used, client_name, impl, hierarchy,
                                   client_dir, feature_counts)

    # 组装报告
    report = build_report(client_name, bought, impl, used, grid, recs)

    if output_path:
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(report)
            print(f"\n报告已保存: {output_path}")
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"\n报告保存失败: {e}")
    else:
        print(report)

    return grid, recs


def build_report(client_name, bought, impl, used, grid, recs) -> str:
    """组装Markdown报告"""
    hierarchy = load_hierarchy()
    mod_kw_map = build_module_kw_map(hierarchy)

    lines = [
        f"# {client_name} - 功能缺口分析与激活建议",
        f"\n**时间**: 2026-03-20",
        "\n## 功能使用诊断\n",
    ]

    for cls, label in [('A','深度应用'), ('B','激活不足'), ('C','疑似未用'), ('D','潜在需求'), ('E','空白机会')]:
        if grid.get(cls):
            lines.append(f"- **[{cls}] {label}**: {', '.join(grid[cls])}")

    lines.append("\n## 详细分析\n")

    for cls in 'ABCD':
        mods = grid.get(cls, [])
        if not mods:
            continue
        rec_data = recs.get(cls, {})
        raw = rec_data.get('raw', '')
        qdrant = rec_data.get('qdrant', [])

        lines.append(f"### {cls}类\n")
        if raw and not raw.startswith('[') and not raw.startswith('['):
            lines.append(f"{raw}\n")

        if qdrant:
            lines.append(f"**推荐产品功能**（知识库）：\n")
            for item in qdrant[:5]:
                feat = item.get('feature', '')
                usage = item.get('usage', '')
                value = item.get('value', '')
                if feat:
                    lines.append(f"- **{feat}**")
                    if usage:
                        lines.append(f"  - 怎么用：{usage}")
                    if value:
                        lines.append(f"  - 业务价值：{value}")
            lines.append("")

    return '\n'.join(lines)


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('client', help='客户名称（部分匹配）')
    parser.add_argument('--year', type=int, default=2025)
    parser.add_argument('--output', help='输出Markdown文件路径')
    args = parser.parse_args()

    try:
        main(args.client, args.year, args.output)
    except FileNotFoundError as e:
        print(f"错误: {e}")
