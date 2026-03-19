#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
产品知识库导入工具（批量加速版）
用法（从skill根目录运行）：
  python scripts/import_knowledge.py              # 全量导入
  python scripts/import_knowledge.py --manual-only  # 仅用户手册
  python scripts/import_knowledge.py --xlsx-only   # 仅迭代清单
"""

import os
import re
import openpyxl
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from src.qdrant_ops import create_collection, add_points_batch, count_points, COLLECTION_NAME

DOCS_ROOT = r"C:\Users\mingh\client-data\raw\产品功能"
XLSX_FILE = os.path.join(DOCS_ROOT, "甄云SRM产品功能清单.xlsx")

MODULE_MAP = {
    "供应商": "供应商管理", "寻源": "寻源管理", "订单": "采购订单",
    "财务": "财务结算", "大数据": "数据应用", "平台": "平台组件",
    "智能应用": "智能应用", "数据应用": "数据应用",
    "敏捷协同": "敏捷协同", "数智化": "数智化套件",
    "需求与商城": "商城采购", "合同与主数据": "合同与主数据",
}


def infer_module(filename: str) -> str:
    for key, module in MODULE_MAP.items():
        if key in filename:
            return module
    return "其他"


def chunk_markdown(content: str, doc_name: str, max_chars: int = 800) -> list:
    """按标题切分MD，每个chunk限制在max_chars内（保留完整性）"""
    chunks = []
    lines = content.split("\n")
    current_section = ""
    current_text = []
    current_len = 0

    def save(section, text_list):
        if not text_list:
            return
        combined = "\n".join(text_list).strip()
        if len(combined) < 30:
            return
        chunks.append({
            "text": combined,
            "metadata": {
                "source": "用户手册", "module": infer_module(doc_name),
                "type": "标准功能", "version": "v2024",
                "doc_name": doc_name, "section": section
            }
        })

    for line in lines:
        if re.match(r"^#{1,3}\s+", line):
            # 新章节，检查是否超长
            if current_len > max_chars:
                save(current_section, current_text)
            elif current_text:
                save(current_section, current_text)
            current_section = re.sub(r"^#+\s+", "", line).strip()
            current_text = [line]
            current_len = len(line)
        else:
            current_text.append(line)
            current_len += len(line)

    if current_text:
        save(current_section, current_text)
    return chunks


def import_manuals() -> int:
    print(f"\n{'='*50}\n导入用户手册（65份MD）...\n{'='*50}")
    md_files = sorted([
        f for f in os.listdir(DOCS_ROOT)
        if os.path.isfile(os.path.join(DOCS_ROOT, f))
        and f.startswith("甄云SRM用户手册")
    ])
    print(f"找到 {len(md_files)} 份文档，切分中...")

    all_items = []
    file_stats = []
    for filename in md_files:
        filepath = os.path.join(DOCS_ROOT, filename)
        content = ""
        for enc in ("utf-8", "gbk", "utf-16"):
            try:
                with open(filepath, "r", encoding=enc) as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue
        if not content:
            continue
        chunks = chunk_markdown(content, filename)
        for chunk in chunks:
            all_items.append({"text": chunk["text"], "metadata": chunk["metadata"]})
        file_stats.append((filename[:40], len(chunks)))

    print(f"共 {len(all_items)} 个chunk，开始批量导入...")

    # 批量导入（每批50）
    BATCH = 50
    total_added = 0
    for i in range(0, len(all_items), BATCH):
        batch = all_items[i:i+BATCH]
        n = add_points_batch(batch, batch_size=BATCH)
        total_added += n
        done = min(i+BATCH, len(all_items))
        print(f"  进度: {done}/{len(all_items)} ({total_added}条已入库)")

    print(f"\n用户手册导入完成: {total_added} 条")
    for name, cnt in file_stats:
        print(f"  {name}: {cnt} chunks")
    return total_added


def import_xlsx() -> int:
    print(f"\n{'='*50}\n导入迭代清单...\n{'='*50}")
    if not os.path.exists(XLSX_FILE):
        print(f"文件不存在: {XLSX_FILE}")
        return 0
    try:
        wb = openpyxl.load_workbook(XLSX_FILE)
        ws = wb.active
    except Exception as e:
        print(f"读取失败: {e}")
        return 0

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"列名: {headers}")

    # 前向填充：套件和模块列只在分组首行有值
    current_product = ""
    current_suite = ""
    current_module = ""

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rd = dict(zip(headers, row))
        product = rd.get("产品") or ""
        suite = rd.get("套件") or ""
        module = rd.get("模块") or ""
        func_dir = rd.get("功能目录") or ""

        if product: current_product = str(product).strip()
        if suite: current_suite = str(suite).strip()
        if module: current_module = str(module).strip()

        if not func_dir:
            continue
        func_dir = str(func_dir).strip()

        # 每个功能目录项作为一条记录
        text = f"{current_suite} > {current_module} > {func_dir}"
        items.append({
            "text": text,
            "metadata": {
                "source": "迭代清单",
                "module": current_suite,
                "type": "标准功能",
                "version": "v2025Q1",
                "doc_name": os.path.basename(XLSX_FILE),
                "product": current_product,
                "suite": current_suite,
                "module_name": current_module,
                "feature": func_dir,
            }
        })

    if not items:
        print("未提取到有效数据")
        return 0

    print(f"找到 {len(items)} 条功能目录，批量导入...")
    total = add_points_batch(items, batch_size=50)
    print(f"迭代清单导入完成: {total} 条")
    return total


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--manual-only", action="store_true")
    parser.add_argument("--xlsx-only", action="store_true")
    args = parser.parse_args()

    print(f"产品知识库导入工具\nCollection: {COLLECTION_NAME}\nDimension: 3072")
    create_collection()

    before = count_points()
    print(f"导入前已有: {before} 条\n")

    total = 0
    if args.xlsx_only:
        total = import_xlsx()
    elif args.manual_only:
        total = import_manuals()
    else:
        man_total = import_manuals()
        xlsx_total = import_xlsx()
        total = man_total + xlsx_total

    after = count_points()
    print(f"\n{'='*50}")
    print(f"导入完成！本次新增: {total} 条")
    print(f"Collection 共: {after} 条")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
