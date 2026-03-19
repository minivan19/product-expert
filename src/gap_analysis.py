#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
场景2：已用功能 → 缺口分析
输入客户名称，自动从多数据源提取已用模块，与产品功能库对比，输出缺口分析报告
"""

import os
import sys
import re
import json
import openpyxl
import requests
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from qdrant_client import search, count_points, COLLECTION_NAME

# 路径配置
CLIENT_DATA_ROOT = r"C:\Users\mingh\client-data\raw\客户档案"

# LLM配置
LLM_API_KEY = "sk-Pzt8a346e78b733bfead64b269317c033e97cd59abfWoqEt"
LLM_BASE_URL = "https://api.gptsapi.net/v1"
LLM_MODEL = "gpt-3.5-turbo"

SEARCH_TOP_K = 50


def find_client_dir(client_name: str) -> str | None:
    """查找客户目录"""
    if not os.path.isdir(CLIENT_DATA_ROOT):
        return None
    for name in os.listdir(CLIENT_DATA_ROOT):
        if client_name in name and os.path.isdir(os.path.join(CLIENT_DATA_ROOT, name)):
            return os.path.join(CLIENT_DATA_ROOT, name)
    return None


def extract_modules_from_contracts(client_dir: str) -> list[dict]:
    """从合同相关文件提取模块清单"""
    modules = []
    biz_info_dir = os.path.join(client_dir, "商务信息")
    if not os.path.isdir(biz_info_dir):
        return modules

    # 查找合同附件、报价单等
    for filename in os.listdir(biz_info_dir):
        if any(k in filename for k in ["合同", "报价", "方案", "模块"]):
            filepath = os.path.join(biz_info_dir, filename)
            try:
                if filename.endswith(".xlsx"):
                    mods = extract_modules_from_xlsx(filepath)
                elif filename.endswith((".md", ".txt")):
                    mods = extract_modules_from_text(filepath)
                else:
                    continue
                modules.extend(mods)
            except Exception as e:
                print(f"    解析合同文件失败 {filename}: {e}")

    return modules


def extract_modules_from_xlsx(filepath: str) -> list[dict]:
    """从Excel提取模块信息"""
    modules = []
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # 读取表头
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(cell.value)

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            # 尝试找模块名列
            for col_name in ["模块", "产品模块", "功能模块", "模块名称", "系统模块"]:
                if col_name in row_dict and row_dict[col_name]:
                    mod = str(row_dict[col_name]).strip()
                    if mod and len(mod) > 1:
                        modules.append({
                            "name": mod,
                            "source": "合同模块",
                            "file": os.path.basename(filepath)
                        })
    except Exception as e:
        pass
    return modules


def extract_modules_from_text(filepath: str) -> list[dict]:
    """从文本文件提取模块信息"""
    modules = []
    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()

        # 常见模块关键词模式
        module_patterns = [
            r"(供应商管理|寻源管理|采购订单|合同管理|价格管理|供应商协同|招投标|竞价|询价|预算管理|采购申请|商城|结算|付款|质量|风控)",
        ]

        for pattern in module_patterns:
            matches = re.findall(pattern, content)
            for m in matches:
                modules.append({
                    "name": m,
                    "source": "合同文本",
                    "file": os.path.basename(filepath)
                })
    except:
        pass
    return modules


def extract_modules_from_workorders(client_dir: str) -> list[dict]:
    """从运维工单提取高频使用模块"""
    modules = []
    ops_dir = os.path.join(client_dir, "运维工单")
    if not os.path.isdir(ops_dir):
        return modules

    # 查找工单Excel
    for filename in os.listdir(ops_dir):
        if filename.endswith(".xlsx") and any(k in filename for k in ["工单", "运维", "SLA"]):
            filepath = os.path.join(ops_dir, filename)
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                ws = wb.active

                # 读取表头
                headers = []
                for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                    headers.append(cell.value)

                module_col = None
                for i, h in enumerate(headers):
                    if h and any(k in str(h) for k in ["模块", "系统", "产品", "分类"]):
                        module_col = i
                        break

                if module_col is None:
                    # 尝试第2列
                    module_col = 1

                module_count = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    val = row[module_col] if module_col < len(row) else None
                    if val:
                        mod = str(val).strip()
                        if mod and len(mod) > 1:
                            module_count[mod] = module_count.get(mod, 0) + 1

                # 按出现频次排序
                for mod, count in sorted(module_count.items(), key=lambda x: -x[1])[:20]:
                    modules.append({
                        "name": mod,
                        "source": f"工单（{count}次）",
                        "file": filename
                    })
            except Exception as e:
                print(f"    解析工单文件失败 {filename}: {e}")

    return modules


def extract_modules_from_blueprint(client_dir: str) -> list[dict]:
    """从蓝图/用户手册提取实施模块"""
    modules = []
    for filename in os.listdir(client_dir):
        if any(k in filename for k in ["蓝图", "实施", "用户手册", "UM", "Manual"]):
            filepath = os.path.join(client_dir, filename)
            try:
                if filename.endswith(".md"):
                    mods = extract_modules_from_text(filepath)
                    for m in mods:
                        m["source"] = "蓝图"
                    modules.extend(mods)
            except:
                pass
    return modules


def call_llm(messages: list) -> str:
    """调用LLM"""
    url = f"{LLM_BASE_URL}/chat/completions"
    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": LLM_MODEL,
        "messages": messages,
        "temperature": 0.3
    }
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=120)
        if response.status_code == 200:
            return response.json()["choices"][0]["message"]["content"]
        else:
            return f"LLM调用失败: {response.status_code}"
    except Exception as e:
        return f"LLM调用异常: {e}"


def generate_gap_report(client_name: str, used_modules: list[dict],
                         product_features: list, output_path: str = None) -> str:
    """生成缺口分析报告"""
    # 去重
    seen = set()
    unique_used = []
    for m in used_modules:
        if m["name"] not in seen:
            seen.add(m["name"])
            unique_used.append(m)

    # 格式化为文本
    used_text = "\n".join([
        f"- {m['name']}（来源：{m['source']}）" for m in unique_used
    ])

    # 产品功能列表
    prod_text = "\n".join([
        f"- {p[0].get('module', '')}/{p[0].get('type', '')}: {p[0].get('text', '')[:100]}..." 
        for p in product_features[:30]
    ])

    system_prompt = """你是一名SRM产品专家。你需要为客户做功能缺口分析。

分析维度：
1. **已用功能**：客户已经在使用的功能（来自合同/工单/蓝图）
2. **未覆盖需求**：客户购买了但未充分使用的功能（从合同模块 vs 产品功能库推断）
3. **可挖掘功能**：客户有需求但当前合同未包含的功能（基于工单分析推断）
4. **优先级建议**：按业务价值/实施复杂度给出推荐

输出格式：
## 客户名称：XXX
## 数据来源
- 合同模块：N个
- 运维工单：N个（共X条工单）
- 蓝图/手册：N个

## 1. 已用功能概览
列出客户当前主要在用的模块（5-10个）

## 2. 功能缺口分析
### 2.1 未充分使用（已购未用）
### 2.2 可挖掘机会（业务有需求但未采购）

## 3. 优先级建议
推荐3-5个优先级最高的功能缺口，附上推荐理由

## 4. 总结"""

    user_prompt = f"""客户名称：{client_name}

已用模块清单：
{used_text}

产品功能知识库（部分）：
{prod_text}

请基于以上信息，生成完整的缺口分析报告。"""

    report = call_llm([
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt}
    ])

    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(f"# {client_name} - 功能缺口分析报告\n\n")
            f.write(f"**生成时间**: 2026-03-19\n\n")
            f.write(report)
        print(f"报告已保存: {output_path}")

    return report


def main():
    import argparse
    parser = argparse.ArgumentParser(description="客户功能缺口分析")
    parser.add_argument("客户名称", help="客户名称（用于查找目录）")
    parser.add_argument("--output", "-o", help="输出文件路径")
    args = parser.parse_args()

    client_name = args.客户名称

    print(f"\n🔍 客户: {client_name}")
    print(f"{'='*50}")

    # 查找客户目录
    client_dir = find_client_dir(client_name)
    if not client_dir:
        print(f"❌ 未找到客户目录: {CLIENT_DATA_ROOT} 中未找到包含 '{client_name}' 的目录")
        sys.exit(1)

    print(f"✅ 找到客户目录: {client_dir}")

    # 提取各数据源
    print("\n📂 提取合同模块...")
    contract_modules = extract_modules_from_contracts(client_dir)
    print(f"   合同模块: {len(contract_modules)} 个")

    print("\n📂 提取运维工单模块...")
    workorder_modules = extract_modules_from_workorders(client_dir)
    print(f"   工单模块: {len(workorder_modules)} 个")

    print("\n📂 提取蓝图/手册模块...")
    blueprint_modules = extract_modules_from_blueprint(client_dir)
    print(f"   蓝图/手册: {len(blueprint_modules)} 个")

    # 合并
    all_modules = contract_modules + workorder_modules + blueprint_modules

    if not all_modules:
        print("\n⚠️  未提取到任何模块信息，请检查客户目录结构")
        sys.exit(1)

    # 检查知识库
    total = count_points()
    if total == 0:
        print("⚠️  知识库为空，请先运行: python import_knowledge.py")
        sys.exit(1)

    # 检索产品功能
    print(f"\n📚 检索产品功能库（{total} 条记录）...")
    # 用模块名拼接后搜索
    module_names = [m["name"] for m in all_modules[:10]]
    query = " ".join(module_names)
    if len(query) < 10:
        query = "SRM采购管理供应商合同"

    results = search(query, top_k=SEARCH_TOP_K)
    print(f"   检索到 {len(results)} 条相关功能")

    # 生成报告
    print(f"\n🤖 生成缺口分析报告...")
    report = generate_gap_report(client_name, all_modules, results, args.output)

    print(f"\n{'='*60}")
    print("缺口分析结果")
    print(f"{'='*60}")
    print(report)


if __name__ == "__main__":
    main()
