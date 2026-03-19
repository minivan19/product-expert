#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将产品功能文档导入 Qdrant
- 65份 MD 用户手册 → 按章节切分后导入
- 迭代功能清单 xlsx → 解析后导入
"""

import os
import sys
import re
import json
import time
import openpyxl
from pathlib import Path

# 注入父目录以导入 qdrant_client
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from qdrant_ops import (
    create_collection, collection_exists, add_point,
    count_points, COLLECTION_NAME
)

# 文档根路径
DOCS_ROOT = r"C:\Users\mingh\client-data\raw\产品功能"
MANUALS_DIR = DOCS_ROOT  # MD文件直接在根目录下
XLSX_FILE = os.path.join(DOCS_ROOT, "甄云SRM产品功能清单.xlsx")

# 模块分类映射（从目录名/文件名推断）
MODULE_MAP = {
    "供应商": "供应商管理",
    "寻源": "寻源管理",
    "订单": "采购订单",
    "财务": "财务结算",
    "大数据": "数据应用",
    "平台": "平台组件",
    "智能应用": "智能应用",
    "数据应用": "数据应用",
    "敏捷协同": "敏捷协同",
    "数智化": "数智化套件",
    "需求与商城": "商城采购",
    "合同与主数据": "合同与主数据",
}


def infer_module(filename: str) -> str:
    """从文件名推断所属模块"""
    for key, module in MODULE_MAP.items():
        if key in filename:
            return module
    return "其他"


def infer_type_from_source(source: str, doc_name: str, xlsx_row: dict = None) -> str:
    """推断功能类型：标准功能/新增/修正"""
    if source == "用户手册":
        return "标准功能"
    elif source == "迭代清单":
        if xlsx_row:
            # 尝试从xlsx的"类型"列判断
            for col in ["类型", "变更类型", "变更方式"]:
                if col in xlsx_row and xlsx_row[col]:
                    val = str(xlsx_row[col])
                    if "新增" in val:
                        return "新增"
                    elif "修正" in val or "优化" in val:
                        return "修正"
        return "新增"
    return "标准功能"


def infer_version_from_xlsx(xlsx_row: dict = None) -> str:
    """从xlsx推断版本"""
    if xlsx_row:
        for col in ["版本", "迭代版本", "发布版本"]:
            if col in xlsx_row and xlsx_row[col]:
                return str(xlsx_row[col]).strip()
    return "v2025Q1"  # 默认


def chunk_markdown(content: str, doc_name: str, max_chars: int = 800) -> list[dict]:
    """
    将Markdown文档按章节切分为多个chunk
    每个chunk = {text, metadata}
    """
    chunks = []

    # 提取标题（# 开头）
    lines = content.split("\n")
    current_section = ""
    current_text = []

    def save_chunk(section, text):
        if not text:
            return
        combined = "\n".join(text).strip()
        if len(combined) < 50:  # 太短的不要
            return
        chunks.append({
            "text": combined,
            "metadata": {
                "source": "用户手册",
                "module": infer_module(doc_name),
                "type": "标准功能",
                "version": "v2024",
                "doc_name": doc_name,
                "section": section
            }
        })

    for line in lines:
        # 新章节（# 标题）
        if re.match(r"^#{1,3}\s+", line):
            save_chunk(current_section, current_text)
            current_section = re.sub(r"^#+\s+", "", line).strip()
            current_text = [line]
        else:
            current_text.append(line)

    save_chunk(current_section, current_text)
    return chunks


def import_manuals() -> int:
    """导入65份MD用户手册"""
    print(f"\n{'='*50}")
    print("导入用户手册（65份MD）...")
    print(f"{'='*50}")

    if not os.path.isdir(MANUALS_DIR):
        print(f"目录不存在: {MANUALS_DIR}")
        return 0

    md_files = [f for f in os.listdir(MANUALS_DIR)
                if os.path.isfile(os.path.join(MANUALS_DIR, f)) and f.startswith("甄云SRM用户手册")]
    print(f"找到 {len(md_files)} 份MD文档")

    total_chunks = 0
    total_added = 0

    for i, filename in enumerate(sorted(md_files), 1):
        filepath = os.path.join(MANUALS_DIR, filename)
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                content = f.read()
        except UnicodeDecodeError:
            try:
                with open(filepath, "r", encoding="gbk") as f:
                    content = f.read()
            except:
                print(f"  [{i}/{len(md_files)}] 读取失败: {filename}")
                continue

        chunks = chunk_markdown(content, filename)
        total_chunks += len(chunks)

        for chunk in chunks:
            if add_point(chunk["text"], chunk["metadata"]):
                total_added += 1

        print(f"  [{i}/{len(md_files)}] {filename[:40]}: {len(chunks)} chunks -> {total_added}")

    print(f"\n用户手册导入完成: {total_added} 条记录")
    return total_added


def import_iteration_list() -> int:
    """导入迭代功能清单xlsx"""
    print(f"\n{'='*50}")
    print("导入迭代功能清单...")
    print(f"{'='*50}")

    if not os.path.exists(XLSX_FILE):
        print(f"文件不存在: {XLSX_FILE}")
        return 0

    try:
        wb = openpyxl.load_workbook(XLSX_FILE)
        ws = wb.active
    except Exception as e:
        print(f"读取xlsx失败: {e}")
        return 0

    # 读取表头
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value)

    print(f"表头: {headers}")

    total_added = 0
    rows_read = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        rows_read += 1
        if not row[0]:  # 跳过空行
            continue

        row_dict = dict(zip(headers, row))

        # 提取功能名称和描述
        func_name = str(row_dict.get("功能名称", "") or row_dict.get("功能点", "") or "").strip()
        func_desc = str(row_dict.get("功能描述", "") or row_dict.get("说明", "") or "").strip()
        module_name = str(row_dict.get("所属模块", "") or row_dict.get("模块", "") or "").strip()

        if not func_name:
            continue

        text = f"{func_name}：{func_desc}" if func_desc else func_name

        # 确定metadata
        func_type = infer_type_from_source("迭代清单", XLSX_FILE, row_dict)
        version = infer_version_from_xlsx(row_dict)
        module = module_name if module_name else infer_module(XLSX_FILE)

        metadata = {
            "source": "迭代清单",
            "module": module,
            "type": func_type,
            "version": version,
            "doc_name": os.path.basename(XLSX_FILE)
        }

        if add_point(text, metadata):
            total_added += 1

    print(f"迭代清单导入完成: {total_added} 条记录（共读 {rows_read} 行）")
    return total_added


def main():
    import argparse
    parser = argparse.ArgumentParser(description="导入产品功能文档到Qdrant")
    parser.add_argument("--incremental", action="store_true", help="增量模式（暂未实现）")
    parser.add_argument("--manual-only", action="store_true", help="仅导入用户手册")
    parser.add_argument("--xlsx-only", action="store_true", help="仅导入迭代清单")
    args = parser.parse_args()

    print(f"\n产品知识库导入工具")
    print(f"目标 Collection: {COLLECTION_NAME}")
    print(f"文档路径: {DOCS_ROOT}")

    # 创建collection（如不存在）
    create_collection()

    # 导入数据
    if args.manual_only:
        total = import_manuals()
    elif args.xlsx_only:
        total = import_iteration_list()
    else:
        manual_count = import_manuals()
        xlsx_count = import_iteration_list()
        total = manual_count + xlsx_count

    print(f"\n{'='*50}")
    print(f"导入完成！总计 {total} 条记录")
    print(f"Collection 中现有 {count_points()} 条记录")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
